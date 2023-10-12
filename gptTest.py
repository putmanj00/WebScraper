from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# URL of the website to scrape
url = "https://en.seedfinder.eu/database/strains/alphabetical/"

# Alphabetical list of pages for strains
# strainAlphabeticalList = ["", "1234567890", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "n", "o", "p", "q", "r", "s", "t", "u", "v", "w", "x", "y", "z"]
strainAlphabeticalList = ["x", "b"]

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write headers to the Excel sheet
sheet["A1"] = "Strain"
sheet["B1"] = "Breeder"
sheet["C1"] = "Indica or Sativa"
sheet["D1"] = "Indoor or Outdoor"
sheet["E1"] = "Flowering Time (Days)"
sheet["F1"] = "Female Seeds?"
sheet["G1"] = "Description"

# Functions
def scrape_strain_description(strain_name, breeder):
    # Construct the URL based on strain name and breeder
    url = f"https://en.seedfinder.eu/strain-info/{strain_name.replace(' ', '_')}/{breeder.replace(' ', '_')}/"

    # Fetch the webpage
    strain_page = requests.get(url).text
    strain_soup = BeautifulSoup(strain_page, "html.parser")

    # Find the description
    table = strain_soup.find('div', class_='partInnerDiv')
    strain_description = ""
    if table:
        for p in table.find_all('p'):
            strain_description += p.text.strip() + " "
        strain_description = strain_description.strip()

    return strain_description

# Set up Selenium webdriver
options = webdriver.ChromeOptions()
options.add_argument("--headless")  # Run Chrome in headless mode (no GUI)
driver = webdriver.Chrome(options=options)

for x in strainAlphabeticalList:
    newUrl = url + x + "/"
    driver.get(newUrl)

    # Wait for the page to load
    time.sleep(5)  # Adjust the wait time as needed

    # Find all rows with strain data
    strain_rows = driver.find_elements(By.CSS_SELECTOR, "table#cannabis-strain-table tbody tr")


    # Loop through each row of strain data
    for row in strain_rows:
        strain = row.find_element(By.CSS_SELECTOR, "th.xs1 a").text
        breeder = row.find_element(By.CSS_SELECTOR, "td.graukleinX.rechts.nowrap").text
        indicaSativa = row.find_element(By.CSS_SELECTOR, "td img[width='20']").get_attribute("title")
        indoorOutdoor = row.find_element(By.CSS_SELECTOR, "td.x20 img[height='14']").get_attribute("title")
        floweringTime = row.find_element(By.CSS_SELECTOR, "td.graukleinX span").text
        femaleSeeds = row.find_element(By.CSS_SELECTOR, "td img[width='12']").get_attribute("title")
        description = scrape_strain_description(strain, breeder)
        
        # Find the first empty row in the Excel sheet
        row_num = sheet.max_row + 1
        
        # Write data to the Excel sheet
        sheet[f"A{row_num}"] = strain
        sheet[f"B{row_num}"] = breeder
        sheet[f"C{row_num}"] = indicaSativa
        sheet[f"D{row_num}"] = indoorOutdoor
        sheet[f"E{row_num}"] = floweringTime
        sheet[f"F{row_num}"] = femaleSeeds
        sheet[f"G{row_num}"] = description
        

    # Save the Excel workbook
    workbook.save("new_strain_data.xlsx")

    # Close the browser
driver.quit()

print("Data has been scraped and exported to strain_data.xlsx")
