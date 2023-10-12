import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook

# Fetch the webpage
strain_page = requests.get("https://en.seedfinder.eu/strain-info/XL_Italian_Shoes/Cult_Classics_Seeds/").text
strain_soup = BeautifulSoup(strain_page, "html.parser")

# Find the description
table = strain_soup.find('div', class_='partInnerDiv')
strain_description = ""
if table:
    for p in table.find_all('p'):
        strain_description += p.text.strip() + " "
    strain_description = strain_description.strip()

# Create a DataFrame
data = {'Strain Description': [strain_description]}
df = pd.DataFrame(data)

# Save to an Excel file
excel_file = "strain_description2.xlsx"
df.to_excel(excel_file, index=False)

print(f'Strain description saved to {excel_file}')

def scrape_strain_description(strain_name, breeder):
    # Construct the URL based on strain name and breeder
    url = f"https://en.seedfinder.eu/strain-info/{strain_name.replace(' ', '_')}/{breeder.replace(' ', '_')}/"

    # Fetch the webpage
    strain_page = requests.get(url).text
    strain_soup = BeautifulSoup(strain_page, "html.parser")

    # Find the description
    table = strain_soup.find('div', class_='partInnerDiv')
    strain_description = ""
    if table:
        for p in table.find_all('p'):
            strain_description += p.text.strip() + " "
        strain_description = strain_description.strip()

    return strain_description

# Example usage:
strain_name = "XL Italian Shoes"
breeder = "Cult Classics Seeds"
description = scrape_strain_description(strain_name, breeder)

# Load the existing Excel workbook
excel_file = "strain_data.xlsx"
workbook = load_workbook(excel_file)
sheet = workbook.active

# Find the column index for "Strain"
column_index = None
for cell in sheet[1]:
    if cell.value == "Strain":
        column_index = cell.column_letter
        break

if column_index:
    # Find the row corresponding to the given strain and breeder
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        if row[0] == strain_name and row[1] == breeder:
            # Get the row number
            row_num = row[0].row

            # Write the description to the corresponding cell in the "Strain Description" column
            sheet[f"Z{row_num}"] = description  # Assuming "Z" is an empty column for descriptions

            # Save the updated Excel workbook
            workbook.save(excel_file)

            print(f'Description added to {strain_name} by {breeder} in {excel_file}')
            break
    else:
        print(f'Strain {strain_name} by {breeder} not found in {excel_file}')
else:
    print("Column 'Strain' not found in the Excel sheet.")
